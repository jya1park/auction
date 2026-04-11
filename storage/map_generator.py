# -*- coding: utf-8 -*-
"""
경매 데이터 지도 시각화 모듈 (카카오맵)

output/ 디렉토리의 최신 xlsx 파일을 읽어 카카오 REST API로 지오코딩한 뒤,
좌표가 포함된 데이터를 JSON으로 삽입한 단일 HTML 파일을 생성합니다.
"""
import os
import json
import glob
import time
import urllib.request
import urllib.parse
import urllib.error
from typing import List, Dict, Optional, Tuple

import config


KAKAO_JS_APP_KEY = "614ddc420a052c47f1b0a7eb2169d862"
KAKAO_REST_API_KEY = getattr(config, "KAKAO_REST_API_KEY", "62a90e550520f584844f36e96c9c0d40")

# 주소 컬럼 후보 (우선순위 순)
ADDRESS_COLUMNS = ["물건주소", "소재지", "주소", "물건소재지", "소재지 및 내역"]

# HTML에 포함할 데이터 컬럼 후보 (있는 것만)
DATA_COLUMNS = [
    "사건번호", "법원", "물건번호",
    "물건주소", "소재지", "주소", "물건소재지", "소재지 및 내역",
    "용도", "감정평가액", "감정가",
    "최저입찰가_표시", "최저입찰가", "최저매각가",
    "입찰기일", "매각기일", "진행상태", "상태",
    "유찰횟수", "입찰방법",
]

# 매각결과 시트 컬럼
RESULT_DATA_COLUMNS = [
    "사건번호", "법원", "물건번호",
    "소재지 및 내역",
    "용도", "감정평가액",
    "매각결과", "매각금액",
    "담당계매각기일(입찰기간)",
]

# 기본 중심 좌표 (수원시)
DEFAULT_LAT = 37.2636
DEFAULT_LNG = 127.0286
DEFAULT_LEVEL = 7

# 샘플 데이터 (use_sample=True 시 사용)
SAMPLE_DATA = [
    {
        "사건번호": "2024타경10001",
        "물건번호": "1",
        "물건주소": "경기도 수원시 팔달구 효원로 1",
        "용도": "아파트",
        "감정평가액": "500,000,000",
        "최저입찰가_표시": "350,000,000(70%)",
        "입찰기일": "2024-03-15",
        "진행상태": "진행중",
    },
    {
        "사건번호": "2024타경10002",
        "물건번호": "1",
        "물건주소": "경기도 수원시 영통구 영통로 234",
        "용도": "아파트",
        "감정평가액": "450,000,000",
        "최저입찰가_표시": "315,000,000(70%)",
        "입찰기일": "2024-03-22",
        "진행상태": "진행중",
    },
    {
        "사건번호": "2024타경10003",
        "물건번호": "1",
        "물건주소": "경기도 수원시 권선구 권선로 921",
        "용도": "아파트",
        "감정평가액": "380,000,000",
        "최저입찰가_표시": "266,000,000(70%)",
        "입찰기일": "2024-04-05",
        "진행상태": "유찰",
    },
    {
        "사건번호": "2024타경10004",
        "물건번호": "1",
        "물건주소": "경기도 용인시 기흥구 구성로 357",
        "용도": "근린시설",
        "감정평가액": "980,000,000",
        "최저입찰가_표시": "686,000,000(70%)",
        "입찰기일": "2024-04-14",
        "진행상태": "진행중",
    },
    {
        "사건번호": "2024타경10005",
        "물건번호": "1",
        "물건주소": "경기도 화성시 동탄반석로 142",
        "용도": "아파트",
        "감정평가액": "620,000,000",
        "최저입찰가_표시": "434,000,000(70%)",
        "입찰기일": "2024-04-19",
        "진행상태": "진행중",
    },
]


# ──────────────────────────────────────────────
# 파일 유틸리티
# ──────────────────────────────────────────────

def _get_output_dir() -> str:
    return os.path.join(os.path.dirname(os.path.dirname(__file__)), config.OUTPUT_DIR)


def _get_geocode_cache_path() -> str:
    return os.path.join(_get_output_dir(), "geocode_cache.json")


def _find_latest_xlsx() -> str:
    """output/ 디렉토리에서 가장 최근 xlsx 파일을 반환합니다 (임시파일 제외)."""
    out_dir = _get_output_dir()
    candidates = [
        f for f in glob.glob(os.path.join(out_dir, "*.xlsx"))
        if not os.path.basename(f).startswith("~$")
    ]
    if not candidates:
        return os.path.join(out_dir, "courtauction_data.xlsx")
    return max(candidates, key=os.path.getmtime)


def _read_excel(xlsx_path: str) -> Tuple[List[str], List[Dict]]:
    """경매목록 시트에서 헤더와 데이터를 읽어 반환합니다."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl이 설치되지 않았습니다. pip install openpyxl")

    print(f"[MapGen] openpyxl로 파일 열기: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "경매목록" in wb.sheetnames:
        ws = wb["경매목록"]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], []

    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        record = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
        data.append(record)

    return headers, data


def _read_result_excel(xlsx_path: str) -> Tuple[List[str], List[Dict]]:
    """매각결과 시트에서 헤더와 데이터를 읽어 반환합니다."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl이 설치되지 않았습니다. pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "매각결과" not in wb.sheetnames:
        wb.close()
        return [], []

    ws = wb["매각결과"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], []

    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        record = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
        data.append(record)

    return headers, data


def _find_address_column(headers: List[str]) -> Optional[str]:
    for candidate in ADDRESS_COLUMNS:
        if candidate in headers:
            return candidate
    for h in headers:
        if "소재" in h or "주소" in h:
            return h
    return None


# ──────────────────────────────────────────────
# 지오코딩 (Kakao REST API + 캐시)
# ──────────────────────────────────────────────

def _load_geocode_cache() -> Dict:
    path = _get_geocode_cache_path()
    if os.path.exists(path):
        for enc in ("utf-8", "utf-8-sig", "cp949"):
            try:
                with open(path, "r", encoding=enc) as f:
                    return json.load(f)
            except UnicodeDecodeError:
                continue
            except Exception:
                break
        # 캐시 파일이 손상됐으면 삭제 후 빈 캐시로 시작
        print(f"[MapGen] 캐시 파일 손상, 초기화: {path}")
        try:
            os.remove(path)
        except Exception:
            pass
    return {}


def _save_geocode_cache(cache: Dict) -> None:
    path = _get_geocode_cache_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def _geocode_address(address: str) -> Optional[Tuple[float, float]]:
    """Kakao REST API로 주소를 (lat, lng)로 변환합니다."""
    url = (
        "https://dapi.kakao.com/v2/local/search/address.json?query="
        + urllib.parse.quote(address)
    )
    req = urllib.request.Request(
        url, headers={"Authorization": f"KakaoAK {KAKAO_REST_API_KEY}"}
    )
    try:
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            docs = data.get("documents", [])
            if docs:
                return float(docs[0]["y"]), float(docs[0]["x"])
    except Exception as e:
        print(f"[MapGen] 지오코딩 실패 ({address[:40]}): {e}")
    return None


def _geocode_items(data: List[Dict], addr_col: str) -> List[Dict]:
    """데이터 목록의 주소를 지오코딩하여 lat/lng 필드를 추가합니다."""
    cache = _load_geocode_cache()
    result = []
    changed = False
    total = len(data)

    for i, item in enumerate(data):
        address = str(item.get(addr_col, "") or "").strip()
        if not address or address == "None":
            continue

        new_item = dict(item)

        if address in cache:
            coords = cache[address]
        else:
            print(f"[MapGen] 지오코딩 ({i + 1}/{total}): {address[:50]}")
            coords = _geocode_address(address)
            if coords is not None:        # 성공한 결과만 캐시 (None은 저장 안 함 → 다음 실행 시 재시도)
                cache[address] = coords
                changed = True
            time.sleep(0.15)  # API 부하 분산

        if coords:
            new_item["lat"] = coords[0]
            new_item["lng"] = coords[1]

        result.append(new_item)

    if changed:
        _save_geocode_cache(cache)
        success = sum(1 for it in result if "lat" in it)
        print(f"[MapGen] 지오코딩 완료: {success}/{len(result)}건 성공, 캐시 저장")

    return result


# ──────────────────────────────────────────────
# JSON 빌더
# ──────────────────────────────────────────────

def _build_items_json(data: List[Dict], headers: List[str], addr_col: str) -> str:
    """HTML에 인라인으로 삽입할 JS 배열 문자열을 반환합니다."""
    items = []
    for record in data:
        address = str(record.get(addr_col, "") or "").strip()
        if not address or address == "None":
            continue

        item: Dict = {"addr": address}

        # Python geocoding 결과
        if record.get("lat") is not None:
            item["lat"] = record["lat"]
            item["lng"] = record["lng"]

        for col in DATA_COLUMNS:
            if col in headers and record.get(col) is not None:
                val = str(record[col]).strip()
                if val and val != "None":
                    item[col] = val

        items.append(item)

    return json.dumps(items, ensure_ascii=False)


def _build_result_items_json(data: List[Dict], headers: List[str], addr_col: str) -> str:
    """매각결과 데이터를 HTML 인라인 JS 배열 문자열로 반환합니다."""
    items = []
    for record in data:
        address = str(record.get(addr_col, "") or "").strip()
        if not address or address == "None":
            continue

        item: Dict = {"addr": address}

        if record.get("lat") is not None:
            item["lat"] = record["lat"]
            item["lng"] = record["lng"]

        for col in RESULT_DATA_COLUMNS:
            if col in headers and record.get(col) is not None:
                val = str(record[col]).strip()
                if val and val != "None":
                    item[col] = val

        items.append(item)

    return json.dumps(items, ensure_ascii=False)


# ──────────────────────────────────────────────
# HTML 빌더
# ──────────────────────────────────────────────

def _build_html(items_json: str, total: int,
                result_items_json: str = "[]", result_total: int = 0,
                title_suffix: str = "") -> str:
    title = f"경매 물건 지도{title_suffix}"

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{title}</title>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: 'Malgun Gothic', sans-serif; }}
    #map {{ width: 100%; height: 100vh; }}

    /* 상단 정보 패널 */
    #panel {{
      position: absolute; top: 10px; left: 10px; z-index: 2;
      background: rgba(255,255,255,0.93);
      padding: 8px 14px; border-radius: 6px;
      box-shadow: 0 2px 8px rgba(0,0,0,0.22);
      font-size: 13px; font-weight: bold; color: #1F4E79;
    }}

    /* 필터 패널 */
    #filter-wrap {{
      position: absolute; top: 48px; left: 10px; z-index: 2;
      display: flex; flex-wrap: wrap; gap: 5px; max-width: 340px;
    }}
    .filter-btn {{
      padding: 4px 10px; border-radius: 14px; border: 2px solid #ccc;
      background: #fff; font-size: 12px; cursor: pointer;
      font-family: 'Malgun Gothic', sans-serif;
      transition: background 0.15s, color 0.15s;
    }}
    .filter-btn.active {{
      color: #fff; border-color: transparent;
    }}

    /* 범례 */
    #legend {{
      position: absolute; bottom: 30px; left: 10px; z-index: 2;
      background: rgba(255,255,255,0.93);
      padding: 8px 12px; border-radius: 6px;
      box-shadow: 0 2px 8px rgba(0,0,0,0.18);
      font-size: 12px;
    }}
    .legend-item {{ display: flex; align-items: center; margin: 3px 0; }}
    .legend-dot {{
      width: 12px; height: 12px; border-radius: 50%; margin-right: 7px; flex-shrink: 0;
    }}

    /* 팝업 (InfoWindow) */
    .iw-wrap {{
      padding: 10px 30px 10px 12px;
      font-family: 'Malgun Gothic', sans-serif; font-size: 13px;
      min-width: 200px; max-width: 300px; position: relative;
    }}
    .iw-close {{
      position: absolute; top: 6px; right: 8px;
      background: none; border: none; font-size: 15px;
      cursor: pointer; color: #888; line-height: 1; padding: 2px;
    }}
    .iw-close:hover {{ color: #333; }}
    .iw-table {{ border-collapse: collapse; width: 100%; margin-top: 4px; }}
    .iw-table th {{
      text-align: left; padding: 2px 10px 2px 0;
      color: #666; white-space: nowrap; vertical-align: top; font-weight: normal;
    }}
    .iw-table td {{ padding: 2px 0; color: #222; word-break: break-all; }}
    .iw-btn {{
      display: inline-block; margin-top: 8px; padding: 4px 10px;
      background: #1F4E79; color: #fff; border: none; border-radius: 4px;
      font-size: 12px; cursor: pointer; font-family: 'Malgun Gothic', sans-serif;
    }}
    .iw-btn:hover {{ background: #2e6ca8; }}

    /* Toast */
    #toast {{
      position: fixed; bottom: 60px; left: 50%; transform: translateX(-50%);
      background: rgba(30,30,30,0.85); color: #fff;
      padding: 9px 20px; border-radius: 20px; font-size: 13px;
      z-index: 9999; opacity: 0; pointer-events: none;
      transition: opacity 0.3s;
    }}
    #toast.show {{ opacity: 1; }}
  </style>
</head>
<body>
  <div id="map"></div>
  <div id="panel">
    경매목록 {total}건 | 매각결과 {result_total}건 | 지도: <span id="mapped-count">0</span>건
  </div>
  <div id="filter-wrap"></div>
  <div id="legend"></div>
  <div id="toast"></div>

  <script>
    var ITEMS = {items_json};
    var RESULT_ITEMS = {result_items_json};
  </script>
  <script src="//dapi.kakao.com/v2/maps/sdk.js?appkey={KAKAO_JS_APP_KEY}&libraries=services,clusterer"></script>
  <script>
    // ── 색상 정의 ──────────────────────────────
    var TYPE_COLORS = {{
      '아파트':    '#E74C3C',
      '오피스텔':  '#9B59B6',
      '근린시설':  '#3498DB',
      '상가':      '#F39C12',
      '토지':      '#27AE60',
      '다세대':    '#1ABC9C',
      '단독주택':  '#795548',
      '기타':      '#7F8C8D',
    }};
    function getColor(type) {{
      return TYPE_COLORS[type] || TYPE_COLORS['기타'];
    }}

    // ── MarkerImage (SVG 핀) ───────────────────
    function makeMarkerImage(color) {{
      var svg = '<svg xmlns="http://www.w3.org/2000/svg" width="22" height="30" viewBox="0 0 22 30">'
        + '<path d="M11 0C4.9 0 0 4.9 0 11c0 8.3 11 19 11 19S22 19.3 22 11C22 4.9 17.1 0 11 0z"'
        + ' fill="' + color + '" stroke="rgba(0,0,0,0.25)" stroke-width="1"/>'
        + '<circle cx="11" cy="11" r="4.5" fill="rgba(255,255,255,0.85)"/>'
        + '</svg>';
      var url = 'data:image/svg+xml;charset=utf-8,' + encodeURIComponent(svg);
      return new kakao.maps.MarkerImage(
        url,
        new kakao.maps.Size(22, 30),
        {{ offset: new kakao.maps.Point(11, 30) }}
      );
    }}

    // ── 매각결과 마커: 낙찰=파란별★, 유찰=주황×핀 ────
    function makeResultMarkerImage(resultType) {{
      var svg;
      if (resultType === '낙찰') {{
        // 파란 핀 + 금색 별★
        svg = '<svg xmlns="http://www.w3.org/2000/svg" width="26" height="34" viewBox="0 0 26 34">'
          + '<path d="M13 0C5.8 0 0 5.8 0 13c0 9.7 13 21 13 21S26 22.7 26 13C26 5.8 20.2 0 13 0z"'
          + ' fill="#1565C0" stroke="rgba(0,0,80,0.35)" stroke-width="1.5"/>'
          + '<polygon points="13,4.5 14.9,10.3 21,10.3 16.1,13.9 17.9,19.7 13,16.1 8.1,19.7 9.9,13.9 5,10.3 11.1,10.3"'
          + ' fill="#FFD700"/>'
          + '</svg>';
      }} else {{
        // 주황 다이아몬드 핀 + 흰색 × 기호
        svg = '<svg xmlns="http://www.w3.org/2000/svg" width="26" height="34" viewBox="0 0 26 34">'
          + '<path d="M13 0C5.8 0 0 5.8 0 13c0 9.7 13 21 13 21S26 22.7 26 13C26 5.8 20.2 0 13 0z"'
          + ' fill="#E65100" stroke="rgba(80,20,0,0.35)" stroke-width="1.5"/>'
          + '<line x1="8" y1="8" x2="18" y2="18" stroke="white" stroke-width="2.5" stroke-linecap="round"/>'
          + '<line x1="18" y1="8" x2="8" y2="18" stroke="white" stroke-width="2.5" stroke-linecap="round"/>'
          + '</svg>';
      }}
      var url = 'data:image/svg+xml;charset=utf-8,' + encodeURIComponent(svg);
      return new kakao.maps.MarkerImage(
        url,
        new kakao.maps.Size(26, 34),
        {{ offset: new kakao.maps.Point(13, 34) }}
      );
    }}

    // ── Toast ──────────────────────────────────
    var _toastTimer = null;
    function showToast(msg) {{
      var el = document.getElementById('toast');
      el.textContent = msg;
      el.classList.add('show');
      clearTimeout(_toastTimer);
      _toastTimer = setTimeout(function() {{ el.classList.remove('show'); }}, 2500);
    }}

    // ── openCase ──────────────────────────────
    function openCase(caseNum, court) {{
      if (!caseNum) return;
      var hash = 'caseNo=' + encodeURIComponent(caseNum);
      if (court) hash += '&court=' + encodeURIComponent(court);
      var url = 'https://www.courtauction.go.kr/pgj/index.on'
              + '?w2xPath=/pgj/ui/pgj100/PGJ159M00.xml#' + hash;
      window.open(url, '_blank');
      showToast(caseNum + ' — 새 탭에서 자동검색 시도 중…');
    }}

    // ── InfoWindow 닫기 ────────────────────────
    var currentIW = null;
    window.closeIW = function() {{
      if (currentIW) {{ currentIW.close(); currentIW = null; }}
    }};

    // ── 팝업 내용 빌더 (경매목록) ─────────────
    var POPUP_LABELS = {{
      '사건번호': '사건번호', '법원': '법원', '물건번호': '물건번호',
      '물건주소': '주소', '소재지': '주소', '주소': '주소', '물건소재지': '주소',
      '용도': '용도',
      '감정평가액': '감정가', '감정가': '감정가',
      '최저입찰가_표시': '최저매각가', '최저입찰가': '최저매각가', '최저매각가': '최저매각가',
      '입찰기일': '매각기일', '매각기일': '매각기일',
      '진행상태': '상태', '상태': '상태',
      '유찰횟수': '유찰횟수',
    }};
    function buildPopupContent(item) {{
      var seen = {{}};
      var rows = '';
      Object.keys(POPUP_LABELS).forEach(function(key) {{
        var label = POPUP_LABELS[key];
        if (item[key] && !seen[label]) {{
          seen[label] = true;
          rows += '<tr><th>' + label + '</th><td>' + item[key] + '</td></tr>';
        }}
      }});
      var caseNum = item['사건번호'] || '';
      var court   = item['법원'] || '';
      var btn = caseNum
        ? '<button class="iw-btn"'
            + ' data-case="' + caseNum.replace(/"/g, '&quot;') + '"'
            + ' data-court="' + court.replace(/"/g, '&quot;') + '"'
            + ' onclick="openCase(this.dataset.case,this.dataset.court)">법원경매 바로가기</button>'
        : '';
      return '<div class="iw-wrap">'
        + '<button class="iw-close" onclick="closeIW()">✕</button>'
        + '<table class="iw-table">' + rows + '</table>'
        + btn
        + '</div>';
    }}

    // ── 팝업 내용 빌더 (매각결과) ─────────────
    var RESULT_POPUP_LABELS = {{
      '사건번호': '사건번호', '법원': '법원', '물건번호': '물건번호',
      '소재지 및 내역': '주소',
      '용도': '용도',
      '감정평가액': '감정가',
      '매각결과': '매각결과',
      '매각금액': '낙찰금액',
      '담당계매각기일(입찰기간)': '매각기일',
    }};
    function buildResultPopupContent(item) {{
      var seen = {{}};
      var rows = '';
      Object.keys(RESULT_POPUP_LABELS).forEach(function(key) {{
        var label = RESULT_POPUP_LABELS[key];
        if (item[key] && !seen[label]) {{
          seen[label] = true;
          var val = item[key];
          if (key === '매각결과') {{
            var color = (val === '낙찰') ? '#1565C0' : '#E65100';
            val = '<b style="color:' + color + '">' + val + '</b>';
          }}
          rows += '<tr><th>' + label + '</th><td>' + val + '</td></tr>';
        }}
      }});
      var headerColor = item['매각결과'] === '낙찰' ? '#1565C0' : '#E65100';
      var headerText  = item['매각결과'] === '낙찰' ? '🏆 낙찰 결과' : '❌ 유찰';
      var caseNum = item['사건번호'] || '';
      var court   = item['법원'] || '';
      var btn = caseNum
        ? '<button class="iw-btn" style="background:' + headerColor + '"'
            + ' data-case="' + caseNum.replace(/"/g, '&quot;') + '"'
            + ' data-court="' + court.replace(/"/g, '&quot;') + '"'
            + ' onclick="openCase(this.dataset.case,this.dataset.court)">법원경매 바로가기</button>'
        : '';
      return '<div class="iw-wrap">'
        + '<button class="iw-close" onclick="closeIW()">✕</button>'
        + '<div style="font-weight:bold;color:' + headerColor + ';margin-bottom:5px;font-size:12px;">' + headerText + '</div>'
        + '<table class="iw-table">' + rows + '</table>'
        + btn
        + '</div>';
    }}

    // ── 메인 ──────────────────────────────────
    window.onload = function() {{
      if (typeof kakao === 'undefined' || !kakao.maps) {{
        document.getElementById('map').innerHTML =
          '<div style="display:flex;flex-direction:column;align-items:center;justify-content:center;height:100%;gap:16px;font-family:sans-serif;">'
          + '<div style="font-size:48px;">🗺️</div>'
          + '<div style="font-size:18px;font-weight:bold;color:#333;">카카오맵 SDK 로드 실패</div>'
          + '<div style="font-size:14px;color:#666;text-align:center;line-height:1.6;">'
          + '카카오 개발자 콘솔에서 도메인을 등록해야 합니다.<br>'
          + '<b>developers.kakao.com</b> → 내 애플리케이션 → 플랫폼 → Web<br>'
          + '사이트 도메인에 <b>http://localhost:8080</b> 추가 후 재시도하세요.'
          + '</div>'
          + '</div>';
        return;
      }}
      var mapContainer = document.getElementById('map');
      var map = new kakao.maps.Map(mapContainer, {{
        center: new kakao.maps.LatLng({DEFAULT_LAT}, {DEFAULT_LNG}),
        level: {DEFAULT_LEVEL}
      }});
      map.addControl(new kakao.maps.ZoomControl(),    kakao.maps.ControlPosition.RIGHT);
      map.addControl(new kakao.maps.MapTypeControl(), kakao.maps.ControlPosition.TOPRIGHT);

      var clusterer = new kakao.maps.MarkerClusterer({{
        map: map,
        averageCenter: true,
        minLevel: 4,
        disableClickZoom: false,
      }});

      var bounds     = new kakao.maps.LatLngBounds();
      var allEntries = [];   // {{ marker, type, isResult }}
      var mappedCount = 0;

      // ── 경매목록 마커 생성 (용도별 색상) ──────
      ITEMS.forEach(function(item) {{
        if (item.lat == null || item.lng == null) return;
        var type  = item['용도'] || '기타';
        var color = getColor(type);
        var pos   = new kakao.maps.LatLng(item.lat, item.lng);
        var marker = new kakao.maps.Marker({{
          position: pos,
          image: makeMarkerImage(color),
          zIndex: 1,
        }});
        var iw = new kakao.maps.InfoWindow({{
          content: buildPopupContent(item),
          removable: false,
        }});
        kakao.maps.event.addListener(marker, 'click', function() {{
          closeIW();
          iw.open(map, marker);
          currentIW = iw;
        }});
        bounds.extend(pos);
        mappedCount++;
        allEntries.push({{ marker: marker, type: type, isResult: false }});
      }});

      // ── 매각결과 마커 생성 (파란색 별 핀) ──────
      RESULT_ITEMS.forEach(function(item) {{
        if (item.lat == null || item.lng == null) return;
        var resultType = item['매각결과'] || '매각결과';
        var pos = new kakao.maps.LatLng(item.lat, item.lng);
        var marker = new kakao.maps.Marker({{
          position: pos,
          image: makeResultMarkerImage(resultType),
          zIndex: 10,
        }});
        var iw = new kakao.maps.InfoWindow({{
          content: buildResultPopupContent(item),
          removable: false,
        }});
        kakao.maps.event.addListener(marker, 'click', function() {{
          closeIW();
          iw.open(map, marker);
          currentIW = iw;
        }});
        bounds.extend(pos);
        mappedCount++;
        allEntries.push({{ marker: marker, type: '__result__', isResult: true, resultType: resultType }});
      }});

      document.getElementById('mapped-count').textContent = mappedCount;

      clusterer.addMarkers(allEntries.map(function(e) {{ return e.marker; }}));
      if (mappedCount > 0) {{ map.setBounds(bounds); }}

      // ── 필터: 전체 / 경매 / 낙찰 / 유찰 (멀티 토글) ──────────
      var hasNakchal = allEntries.some(function(e) {{ return e.isResult && e.resultType === '낙찰'; }});
      var hasYuchal  = allEntries.some(function(e) {{ return e.isResult && e.resultType === '유찰'; }});

      // 활성 상태 (true = 표시)
      var vis = {{ auction: true, nakchul: true, yuchal: true }};

      // 마커를 vis 상태에 따라 클러스터러에 반영
      function applyVis() {{
        clusterer.clear();
        closeIW();
        var filtered = allEntries.filter(function(e) {{
          if (!e.isResult)             return vis.auction;
          if (e.resultType === '낙찰') return vis.nakchul;
          if (e.resultType === '유찰') return vis.yuchal;
          return true;
        }});
        clusterer.addMarkers(filtered.map(function(e) {{ return e.marker; }}));
      }}

      // 버튼 스타일 동기화
      var btnAll, btnAuction, btnNak, btnYuc;

      function syncBtnStyle(btn, on, color) {{
        if (on) {{
          btn.style.background  = color;
          btn.style.color       = '#fff';
          btn.style.borderColor = color;
          btn.style.opacity     = '1';
        }} else {{
          btn.style.background  = '#fff';
          btn.style.color       = '#888';
          btn.style.borderColor = '#ccc';
          btn.style.opacity     = '0.7';
        }}
      }}

      function syncAllBtn() {{
        // 전체 버튼: 셋 다 켜져 있으면 활성, 아니면 비활성
        var allOn = vis.auction && vis.nakchul && vis.yuchal;
        syncBtnStyle(btnAll, allOn, '#1F4E79');
      }}

      // ── 버튼 생성 헬퍼 ────────────────────────
      function makeToggleBtn(label, color) {{
        var btn = document.createElement('button');
        btn.className    = 'filter-btn';
        btn.textContent  = label;
        btn.style.borderColor = color;
        btn.style.background  = color;
        btn.style.color       = '#fff';
        return btn;
      }}

      var wrap = document.getElementById('filter-wrap');

      // 전체 버튼
      btnAll = makeToggleBtn('전체', '#1F4E79');
      btnAll.addEventListener('click', function() {{
        // 셋 다 켜져있으면 → 모두 끄기, 아니면 → 모두 켜기
        var allOn = vis.auction && vis.nakchul && vis.yuchal;
        vis.auction = vis.nakchul = vis.yuchal = !allOn;
        syncBtnStyle(btnAuction, vis.auction, '#455A64');
        if (btnNak) syncBtnStyle(btnNak, vis.nakchul, '#1565C0');
        if (btnYuc) syncBtnStyle(btnYuc, vis.yuchal,  '#E65100');
        syncAllBtn();
        applyVis();
        showToast(!allOn ? '전체 표시' : '전체 숨김');
      }});
      wrap.appendChild(btnAll);

      // 경매 버튼
      btnAuction = makeToggleBtn('경매', '#455A64');
      btnAuction.addEventListener('click', function() {{
        vis.auction = !vis.auction;
        syncBtnStyle(btnAuction, vis.auction, '#455A64');
        syncAllBtn();
        applyVis();
        showToast(vis.auction ? '경매목록 표시' : '경매목록 숨김');
      }});
      wrap.appendChild(btnAuction);

      // 낙찰 버튼
      if (hasNakchal) {{
        btnNak = makeToggleBtn('★ 낙찰', '#1565C0');
        btnNak.addEventListener('click', function() {{
          vis.nakchul = !vis.nakchul;
          syncBtnStyle(btnNak, vis.nakchul, '#1565C0');
          syncAllBtn();
          applyVis();
          showToast(vis.nakchul ? '낙찰 표시' : '낙찰 숨김');
        }});
        wrap.appendChild(btnNak);
      }}

      // 유찰 버튼
      if (hasYuchal) {{
        btnYuc = makeToggleBtn('✕ 유찰', '#E65100');
        btnYuc.addEventListener('click', function() {{
          vis.yuchal = !vis.yuchal;
          syncBtnStyle(btnYuc, vis.yuchal, '#E65100');
          syncAllBtn();
          applyVis();
          showToast(vis.yuchal ? '유찰 표시' : '유찰 숨김');
        }});
        wrap.appendChild(btnYuc);
      }}

      // ── 범례 빌더 ────────────────────────────
      var legend = document.getElementById('legend');
      // 경매목록은 용도별 색상이므로 기존 범례 유지
      var listTypes = [];
      allEntries.forEach(function(e) {{
        if (!e.isResult && listTypes.indexOf(e.type) === -1) listTypes.push(e.type);
      }});
      listTypes.forEach(function(type) {{
        var row = document.createElement('div');
        row.className = 'legend-item';
        var dot = document.createElement('div');
        dot.className = 'legend-dot';
        dot.style.background = getColor(type);
        var lbl = document.createElement('span');
        lbl.textContent = type;
        row.appendChild(dot);
        row.appendChild(lbl);
        legend.appendChild(row);
      }});
      // 매각결과 구분선
      if (RESULT_ITEMS.length > 0) {{
        var sep = document.createElement('div');
        sep.style.cssText = 'border-top:1px solid #ddd;margin:5px 0 3px;';
        legend.appendChild(sep);
      }}
      if (hasNakchal) {{
        var rn = document.createElement('div');
        rn.className = 'legend-item';
        rn.innerHTML = '<span style="font-size:13px;color:#1565C0;margin-right:6px;font-weight:bold;">★</span>'
          + '<span style="color:#1565C0;font-weight:bold;">낙찰</span>';
        legend.appendChild(rn);
      }}
      if (hasYuchal) {{
        var ry = document.createElement('div');
        ry.className = 'legend-item';
        ry.innerHTML = '<span style="font-size:13px;color:#E65100;margin-right:6px;font-weight:bold;">✕</span>'
          + '<span style="color:#E65100;font-weight:bold;">유찰</span>';
        legend.appendChild(ry);
      }}
    }};
  </script>
</body>
</html>
"""


# ──────────────────────────────────────────────
# 공개 API
# ──────────────────────────────────────────────

def generate_map(
    xlsx_path: str = None,
    output_path: str = None,
    use_sample: bool = False,
) -> str:
    """
    엑셀 파일을 읽어 카카오맵 HTML을 생성합니다.
    좌표는 Python에서 카카오 REST API로 미리 계산하여 JSON으로 임베드합니다.

    Args:
        xlsx_path:   입력 엑셀 경로 (기본: output/ 최신 xlsx 자동 탐색)
        output_path: 출력 HTML 경로 (기본: output/auction_map.html)
        use_sample:  True이면 엑셀 없이 샘플 더미 데이터로 HTML 생성
    Returns:
        생성된 HTML 파일 경로 (실패 시 빈 문자열)
    """
    if output_path is None:
        output_path = os.path.join(_get_output_dir(), "auction_map.html")

    if use_sample:
        print("[MapGen] 샘플 데이터로 지도 생성 중...")
        data    = SAMPLE_DATA
        headers = list(data[0].keys())
        addr_col = _find_address_column(headers)
        geocoded = _geocode_items(data, addr_col)
        items_json = _build_items_json(geocoded, list(geocoded[0].keys()) if geocoded else headers, addr_col)
        html_content = _build_html(items_json, len(data), title_suffix=" (샘플)")
        label = "샘플 "
    else:
        if xlsx_path is None:
            xlsx_path = _find_latest_xlsx()

        print(f"[MapGen] 엑셀 읽기: {xlsx_path}")
        if not os.path.exists(xlsx_path):
            print(f"[MapGen] 파일 없음: {xlsx_path}")
            return ""

        headers, data = _read_excel(xlsx_path)
        if not data:
            print("[MapGen] 데이터가 없습니다.")
            return ""

        addr_col = _find_address_column(headers)
        if not addr_col:
            print(f"[MapGen] 주소 컬럼을 찾을 수 없습니다. (헤더: {headers})")
            return ""

        print(f"[MapGen] 주소 컬럼: '{addr_col}', {len(data)}건 지오코딩 시작...")
        geocoded    = _geocode_items(data, addr_col)
        all_headers = headers + ["lat", "lng"]
        items_json  = _build_items_json(geocoded, all_headers, addr_col)

        # ── 매각결과 시트 읽기 ──
        result_items_json = "[]"
        result_total = 0
        try:
            r_headers, r_data = _read_result_excel(xlsx_path)
            if r_data:
                r_addr_col = _find_address_column(r_headers)
                if r_addr_col:
                    print(f"[MapGen] 매각결과 주소 컬럼: '{r_addr_col}', {len(r_data)}건 지오코딩...")
                    r_geocoded = _geocode_items(r_data, r_addr_col)
                    r_all_hdrs = r_headers + ["lat", "lng"]
                    result_items_json = _build_result_items_json(r_geocoded, r_all_hdrs, r_addr_col)
                    result_total = len(r_data)
        except Exception as e:
            print(f"[MapGen] 매각결과 시트 처리 오류 (무시): {e}")

        html_content = _build_html(items_json, len(data), result_items_json, result_total)
        label = ""

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"[MapGen] {label}지도 생성 완료: {output_path}")
    return output_path


def upload_to_github(
    filepath: str,
    token: str,
    owner: str = "jya1park",
    repo: str = "auction",
    branch: str = "main",
    remote_dir: str = "storage",
) -> bool:
    """GitHub REST API로 파일을 업로드(또는 갱신)합니다."""
    import base64

    filename    = os.path.basename(filepath)
    remote_path = f"{remote_dir}/{filename}" if remote_dir else filename
    api_url     = f"https://api.github.com/repos/{owner}/{repo}/contents/{remote_path}"

    with open(filepath, "rb") as f:
        content_b64 = base64.b64encode(f.read()).decode("utf-8")

    req_headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github+json",
        "Content-Type": "application/json",
        "X-GitHub-Api-Version": "2022-11-28",
    }

    sha = None
    req_get = urllib.request.Request(api_url, headers=req_headers, method="GET")
    try:
        with urllib.request.urlopen(req_get, timeout=10) as resp:
            sha = json.loads(resp.read().decode("utf-8")).get("sha")
    except urllib.error.HTTPError as e:
        if e.code != 404:
            print(f"[MapGen] GitHub SHA 조회 오류: {e}")
            return False
    except Exception as e:
        print(f"[MapGen] GitHub 연결 오류: {e}")
        return False

    payload: Dict = {
        "message": f"feat: update {filename}",
        "content": content_b64,
        "branch": branch,
    }
    if sha:
        payload["sha"] = sha

    req_put = urllib.request.Request(
        api_url,
        data=json.dumps(payload).encode("utf-8"),
        headers=req_headers,
        method="PUT",
    )
    try:
        with urllib.request.urlopen(req_put, timeout=15) as resp:
            result = json.loads(resp.read().decode("utf-8"))
            html_url = result.get("content", {}).get("html_url", "")
            print(f"[MapGen] GitHub 업로드 완료: {html_url}")
            return True
    except urllib.error.HTTPError as e:
        print(f"[MapGen] GitHub 업로드 실패 ({e.code}): {e.read().decode()}")
        return False
    except Exception as e:
        print(f"[MapGen] GitHub 업로드 오류: {e}")
        return False

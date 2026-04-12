"""
데이터 저장 모듈
CSV와 Excel 형식으로 수집 데이터를 저장합니다.
"""
import os
import csv
from datetime import datetime
from typing import List, Dict

import config

# 고정 Excel 파일명 (항상 이 파일에 누적 업데이트)
EXCEL_FIXED_FILENAME = "courtauction_data.xlsx"


def ensure_output_dir() -> str:
    """출력 디렉토리를 생성하고 경로를 반환합니다."""
    out_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), config.OUTPUT_DIR)
    os.makedirs(out_dir, exist_ok=True)
    return out_dir


def get_timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def get_datetime_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def save_csv(data: List[Dict], filename: str = None) -> str:
    """
    데이터를 CSV 파일로 저장합니다.

    Args:
        data: 저장할 딕셔너리 목록
        filename: 파일명 (없으면 타임스탬프 자동 생성)
    Returns:
        저장된 파일 경로
    """
    if not data:
        print("[Exporter] 저장할 데이터가 없습니다.")
        return ""

    out_dir = ensure_output_dir()
    if not filename:
        filename = f"courtauction_{get_timestamp()}.csv"

    filepath = os.path.join(out_dir, filename)

    # 모든 딕셔너리의 키를 합쳐 헤더 생성 (순서 유지)
    fieldnames = []
    seen = set()
    for row in data:
        for k in row.keys():
            if k not in seen:
                fieldnames.append(k)
                seen.add(k)

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(data)

    print(f"[Exporter] CSV 저장 완료: {filepath} ({len(data)}건)")

    # 지도 생성용 고정명 CSV도 함께 갱신 (generate_map이 읽는 파일)
    list_path = os.path.join(out_dir, "courtauction_list.csv")
    with open(list_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(data)
    print(f"[Exporter] 지도용 CSV 갱신: {list_path}")

    return filepath


def save_result_csv(result_data: List[Dict]) -> str:
    """
    매각결과 데이터를 CSV 파일로 저장합니다.
    항상 고정 파일명 courtauction_result.csv 에 덮어씁니다.
    """
    if not result_data:
        print("[Exporter] 저장할 매각결과 데이터가 없습니다.")
        return ""

    out_dir = ensure_output_dir()
    filepath = os.path.join(out_dir, "courtauction_result.csv")

    fieldnames = []
    seen = set()
    for row in result_data:
        for k in row.keys():
            if k not in seen:
                fieldnames.append(k)
                seen.add(k)

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(result_data)

    print(f"[Exporter] 매각결과 CSV 저장 완료: {filepath} ({len(result_data)}건)")
    return filepath


def save_excel(data: List[Dict], filename: str = None) -> str:
    """
    데이터를 Excel 파일로 저장합니다.
    - 컬럼 너비 자동 조정
    - 금액 컬럼에 만원 단위 추가
    - 헤더 스타일 적용

    Args:
        data: 저장할 딕셔너리 목록
        filename: 파일명 (없으면 타임스탬프 자동 생성)
    Returns:
        저장된 파일 경로
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[Exporter] openpyxl이 설치되지 않았습니다. pip install openpyxl")
        return ""

    if not data:
        print("[Exporter] 저장할 데이터가 없습니다.")
        return ""

    out_dir = ensure_output_dir()
    if not filename:
        filename = f"courtauction_{get_timestamp()}.xlsx"

    filepath = os.path.join(out_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "경매목록"

    # 헤더 구성 (만원 단위 컬럼 추가)
    fieldnames = []
    seen = set()
    for row in data:
        for k in row.keys():
            if k not in seen:
                fieldnames.append(k)
                seen.add(k)

    # 금액 원 컬럼 뒤에 만원 컬럼 삽입
    extended_fields = []
    for f in fieldnames:
        extended_fields.append(f)
        if f.endswith("_원") and f.replace("_원", "") in fieldnames:
            pass  # 이미 _원 컬럼이면 건너뜀
        elif not f.endswith("_원") and f + "_원" not in fieldnames:
            # 금액처럼 보이는 컬럼에 만원 단위 추가 (비율/횟수 컬럼 제외)
            if any(kw in f for kw in ["감정가", "최저입찰가", "낙찰가"]) and not f.endswith("율") and not f.endswith("_표시"):
                extended_fields.append(f + "_만원")

    # 헤더 스타일
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 헤더 행 작성
    for col_idx, field in enumerate(extended_fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 25

    # 데이터 행 작성
    data_font = Font(name="맑은 고딕", size=9)
    data_align = Alignment(vertical="center")
    money_align = Alignment(horizontal="right", vertical="center")
    alt_fill = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")

    for row_idx, row_data in enumerate(data, 2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, field in enumerate(extended_fields, 1):
            # 만원 단위 컬럼 처리
            if field.endswith("_만원"):
                base_field = field.replace("_만원", "") + "_원"
                raw_val = row_data.get(base_field)
                if raw_val is None:
                    base_field2 = field.replace("_만원", "")
                    from crawler.list_parser import parse_amount
                    raw_val = parse_amount(str(row_data.get(base_field2, "")))
                if raw_val:
                    try:
                        value = int(raw_val) // 10_000
                    except (TypeError, ValueError):
                        value = None
                else:
                    value = None
            else:
                value = row_data.get(field)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

            # 금액 컬럼은 오른쪽 정렬
            if any(kw in field for kw in ["_원", "_만원", "감정가", "최저입찰가", "낙찰가"]):
                cell.alignment = money_align
            else:
                cell.alignment = data_align

            if is_alt:
                cell.fill = alt_fill

        ws.row_dimensions[row_idx].height = 18

    # 컬럼 너비 자동 조정
    for col_idx, field in enumerate(extended_fields, 1):
        col_letter = get_column_letter(col_idx)
        # 헤더 너비 계산 (한글 2배)
        header_len = sum(2 if ord(c) > 127 else 1 for c in field)
        # 데이터 최대 너비
        max_data_len = 0
        for row_data in data:
            val = str(row_data.get(field, ""))
            val_len = sum(2 if ord(c) > 127 else 1 for c in val)
            max_data_len = max(max_data_len, val_len)
        col_width = max(header_len + 2, min(max_data_len + 2, 40))
        ws.column_dimensions[col_letter].width = col_width

    # 헤더 행 고정
    ws.freeze_panes = "A2"

    # 자동 필터
    ws.auto_filter.ref = f"A1:{get_column_letter(len(extended_fields))}1"

    wb.save(filepath)
    print(f"[Exporter] Excel 저장 완료: {filepath} ({len(data)}건)")
    return filepath


def _build_sheet(ws, data: List[Dict], sheet_title: str,
                 header_color: str = "1F4E79", date_col_color: str = "2E7D32"):
    """
    공통 시트 빌더: 데이터를 스타일 적용하여 ws에 기록합니다.
    업데이트일시를 첫 번째 컬럼으로, 나머지는 원본 순서대로 배치합니다.
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    ws.title = sheet_title

    # 헤더 구성 (업데이트일시 → 금액 관련 원/만원 컬럼 포함)
    fieldnames = ["업데이트일시"]
    seen = {"업데이트일시"}
    for row in data:
        for k in row.keys():
            if k not in seen:
                fieldnames.append(k)
                seen.add(k)

    extended_fields = []
    for f in fieldnames:
        extended_fields.append(f)
        if not f.endswith("_원") and f + "_원" not in fieldnames:
            if any(kw in f for kw in ["감정가", "최저매각가격", "매각금액", "최저입찰가", "낙찰가"]) \
                    and not f.endswith("율") and not f.endswith("_표시"):
                extended_fields.append(f + "_만원")

    # 스타일
    h_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    d_fill = PatternFill(start_color=date_col_color, end_color=date_col_color, fill_type="solid")
    h_font = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")
    d_font = Font(name="맑은 고딕", size=9)
    c_align = Alignment(vertical="center")
    r_align = Alignment(horizontal="right", vertical="center")

    # 헤더 행
    for ci, field in enumerate(extended_fields, 1):
        cell = ws.cell(row=1, column=ci, value=field)
        cell.fill = d_fill if field == "업데이트일시" else h_fill
        cell.font = h_font
        cell.alignment = h_align
        cell.border = thin
    ws.row_dimensions[1].height = 25

    # 데이터 행
    for ri, row_data in enumerate(data, 2):
        is_alt = (ri % 2 == 0)
        for ci, field in enumerate(extended_fields, 1):
            if field.endswith("_만원"):
                base = field.replace("_만원", "") + "_원"
                raw = row_data.get(base)
                if raw is None:
                    from crawler.list_parser import parse_amount
                    raw = parse_amount(str(row_data.get(field.replace("_만원", ""), "")))
                value = int(raw) // 10_000 if raw else None
            else:
                value = row_data.get(field)
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = d_font
            cell.border = thin
            if any(kw in field for kw in ["_원", "_만원", "감정", "매각금액", "최저", "낙찰"]):
                cell.alignment = r_align
            else:
                cell.alignment = c_align
            if is_alt:
                cell.fill = alt_fill
        ws.row_dimensions[ri].height = 18

    # 컬럼 너비 자동 조정
    for ci, field in enumerate(extended_fields, 1):
        col_letter = get_column_letter(ci)
        h_len = sum(2 if ord(c) > 127 else 1 for c in field)
        max_len = 0
        for row_data in data:
            val = str(row_data.get(field, ""))
            max_len = max(max_len, sum(2 if ord(c) > 127 else 1 for c in val))
        ws.column_dimensions[col_letter].width = max(h_len + 2, min(max_len + 2, 40))

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(extended_fields))}1"


def update_excel(data: List[Dict], result_data: List[Dict] = None) -> str:
    """
    고정 Excel 파일(courtauction_data.xlsx)에 데이터를 누적 업데이트합니다.

    - 시트1 '경매목록': 경매 물건 목록 (사건번호+물건번호 기준 upsert)
    - 시트2 '매각결과': 매각결과 조회 데이터 (사건번호+물건번호 기준 upsert)
    - 모든 행에 '업데이트일시' 열 기록
    - 파일이 없으면 새로 생성, 있으면 기존 데이터와 병합
    """
    try:
        import openpyxl
    except ImportError:
        print("[Exporter] openpyxl이 설치되지 않았습니다. pip install openpyxl")
        return ""

    if not data and not result_data:
        print("[Exporter] 저장할 데이터가 없습니다.")
        return ""

    out_dir = ensure_output_dir()
    filepath = os.path.join(out_dir, EXCEL_FIXED_FILENAME)
    now_str = get_datetime_str()

    def load_sheet_data(wb, sheet_name: str) -> List[Dict]:
        """기존 파일의 특정 시트에서 데이터를 로드합니다."""
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append(dict(zip(headers, row)))
        return rows

    def upsert(existing: List[Dict], new: List[Dict]) -> tuple:
        """사건번호+물건번호 기준 upsert, (merged, new_cnt, upd_cnt) 반환."""
        key_fn = lambda r: (str(r.get("사건번호", "")), str(r.get("물건번호", "")))
        m = {key_fn(r): r for r in existing}
        nc, uc = 0, 0
        for row in new:
            k = key_fn(row)
            if k in m:
                m[k] = row; uc += 1
            else:
                m[k] = row; nc += 1
        return list(m.values()), nc, uc

    # 기존 파일 로드
    existing_list, existing_result = [], []
    if os.path.exists(filepath):
        try:
            wb_old = openpyxl.load_workbook(filepath)
            existing_list = load_sheet_data(wb_old, "경매목록")
            existing_result = load_sheet_data(wb_old, "매각결과")
            print(f"[Exporter] 기존 로드 - 경매목록:{len(existing_list)}건, 매각결과:{len(existing_result)}건")
        except Exception as e:
            print(f"[Exporter] 기존 파일 로드 실패 (새로 생성): {e}")

    # 경매목록 upsert
    merged_list = existing_list
    if data:
        for row in data:
            row["업데이트일시"] = now_str
        merged_list, nc, uc = upsert(existing_list, data)
        print(f"[Exporter] 경매목록: 신규 {nc}건 추가, {uc}건 업데이트 → 총 {len(merged_list)}건")

    # 매각결과 upsert
    merged_result = existing_result
    if result_data:
        for row in result_data:
            row["업데이트일시"] = now_str
        merged_result, nc, uc = upsert(existing_result, result_data)
        print(f"[Exporter] 매각결과: 신규 {nc}건 추가, {uc}건 업데이트 → 총 {len(merged_result)}건")

    # Excel 생성
    wb = openpyxl.Workbook()

    # 시트1: 경매목록 (최신 업데이트 순 정렬)
    ws1 = wb.active
    sorted_list = sorted(merged_list, key=lambda r: str(r.get("업데이트일시", "")), reverse=True)
    _build_sheet(ws1, sorted_list, "경매목록", header_color="1F4E79", date_col_color="2E7D32")

    # 시트2: 매각결과 (최신 업데이트 순 정렬)
    ws2 = wb.create_sheet("매각결과")
    sorted_result = sorted(merged_result, key=lambda r: str(r.get("업데이트일시", "")), reverse=True)
    _build_sheet(ws2, sorted_result, "매각결과", header_color="4A148C", date_col_color="1565C0")

    wb.save(filepath)
    print(f"[Exporter] Excel 업데이트 완료: {filepath}")
    return filepath


def print_summary(data: List[Dict]):
    """수집 데이터의 요약 통계를 출력합니다."""
    if not data:
        print("데이터가 없습니다.")
        return

    print("\n" + "="*60)
    print(f"수집 결과 요약")
    print("="*60)
    print(f"총 수집 건수: {len(data)}건")

    # 감정가 통계
    amounts = [v for v in (d.get("감정가_원") for d in data) if v]
    if amounts:
        print(f"감정가 범위: {min(amounts):,}원 ~ {max(amounts):,}원")
        print(f"감정가 평균: {sum(amounts)//len(amounts):,}원")

    # 최저입찰가 통계
    min_bids = [v for v in (d.get("최저입찰가_원") for d in data) if v]
    if min_bids:
        print(f"최저입찰가 평균: {sum(min_bids)//len(min_bids):,}원")

    # 입찰기일 분포
    dates = {}
    for d in data:
        date = d.get("입찰기일", "")
        if date:
            dates[date] = dates.get(date, 0) + 1
    if dates:
        print(f"\n입찰기일 분포 (상위 5):")
        for date, cnt in sorted(dates.items(), key=lambda x: -x[1])[:5]:
            print(f"  {date}: {cnt}건")

    print("="*60)

# -*- coding: utf-8 -*-
"""
지도 생성 에러 진단 스크립트
python diagnose_map.py 로 실행하면 에러 위치를 정확히 알 수 있습니다.
"""
import os, sys, traceback

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

print("=" * 60)
print("진단 시작")
print("=" * 60)

# 1. config 임포트
print("\n[1] config 임포트...")
try:
    import config
    print(f"    OK: OUTPUT_DIR={config.OUTPUT_DIR}, REST_KEY={getattr(config, 'KAKAO_REST_API_KEY', 'MISSING')[:8]}...")
except Exception:
    print("    FAIL:")
    traceback.print_exc()
    sys.exit(1)

# 2. output 폴더 파일 목록
print("\n[2] output 폴더 파일 목록...")
out_dir = os.path.join(ROOT, config.OUTPUT_DIR)
try:
    files = os.listdir(out_dir)
    for f in files:
        fpath = os.path.join(out_dir, f)
        size = os.path.getsize(fpath)
        with open(fpath, "rb") as fp:
            first4 = fp.read(4).hex()
        print(f"    {f}: {size} bytes, 첫4바이트={first4}")
except Exception:
    traceback.print_exc()

# 3. xlsx 파일 찾기
print("\n[3] xlsx 파일 탐색...")
try:
    import glob
    candidates = [
        f for f in glob.glob(os.path.join(out_dir, "*.xlsx"))
        if not os.path.basename(f).startswith("~$")
    ]
    print(f"    발견된 xlsx: {candidates}")
    if not candidates:
        print("    xlsx 없음 — 샘플 데이터로 테스트합니다")
        xlsx_path = None
    else:
        xlsx_path = max(candidates, key=os.path.getmtime)
        print(f"    사용할 xlsx: {xlsx_path}")
except Exception:
    traceback.print_exc()
    xlsx_path = None

# 4. openpyxl 읽기
if xlsx_path:
    print(f"\n[4] openpyxl 읽기: {xlsx_path}")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        print(f"    시트: {wb.sheetnames}")
        ws = wb["경매목록"] if "경매목록" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        print(f"    행 수: {len(rows)}, 헤더: {rows[0] if rows else '없음'}")
    except Exception:
        print("    FAIL:")
        traceback.print_exc()

# 5. geocode_cache 읽기
print("\n[5] geocode_cache.json 읽기...")
cache_path = os.path.join(out_dir, "geocode_cache.json")
if os.path.exists(cache_path):
    with open(cache_path, "rb") as fp:
        first4 = fp.read(4).hex()
        fp.seek(0)
        size = os.path.getsize(cache_path)
    print(f"    캐시 파일: {size} bytes, 첫4바이트={first4}")
    for enc in ("utf-8", "utf-8-sig", "cp949"):
        try:
            import json
            with open(cache_path, "r", encoding=enc) as f:
                data = json.load(f)
            print(f"    OK ({enc}): {len(data)}개 항목")
            break
        except UnicodeDecodeError:
            print(f"    {enc} 실패")
        except Exception as e:
            print(f"    {enc} 오류: {e}")
else:
    print("    캐시 파일 없음 (정상)")

# 6. map_generator 임포트
print("\n[6] map_generator 임포트...")
try:
    from storage.map_generator import generate_map
    print("    OK")
except Exception:
    print("    FAIL:")
    traceback.print_exc()
    sys.exit(1)

# 7. generate_map 실행 (샘플 데이터)
print("\n[7] generate_map(use_sample=True) 실행...")
try:
    result = generate_map(use_sample=True)
    print(f"    OK: {result}")
except Exception:
    print("    FAIL:")
    traceback.print_exc()

# 8. generate_map 실행 (실제 xlsx)
if xlsx_path:
    print(f"\n[8] generate_map(xlsx_path='{os.path.basename(xlsx_path)}') 실행...")
    try:
        result = generate_map(xlsx_path=xlsx_path)
        print(f"    OK: {result}")
    except Exception:
        print("    FAIL:")
        traceback.print_exc()

print("\n" + "=" * 60)
print("진단 완료")
print("=" * 60)
